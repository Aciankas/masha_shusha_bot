import re
from openpyxl import Workbook
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.dispatcher.filters import Text
from aiogram import types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile

import entities
from create_bot import dp, admin_ids, base, cur, bot_id, delete_message, group_msg
from dtbase import get_slide_deprecated, db_execute, get_keyboard, create_new_button, get_button_by_ids, delete_button_by_ids, \
    insert_media
from message_constructor import construct_slide, SlideError, construct_keyboard_from_select, \
    construct_slide_from_message
from aiogram.utils.exceptions import WrongFileIdentifier


def slide_change_kb(slide_id: int):
    return InlineKeyboardMarkup()\
        .row(InlineKeyboardButton(text='change_media',
                                  callback_data=f'change_media={slide_id}'))\
        .row(InlineKeyboardButton(text='change_message',
                                  callback_data=f'change_message={slide_id}'))\
        .row(InlineKeyboardButton(text='change_bot_id',
                                  callback_data=f'change_bot_id={slide_id}'))\
        .row(InlineKeyboardButton(text='change_modifier',
                                  callback_data=f'change_modifier={slide_id}'))\
        .row(InlineKeyboardButton(text='change_appearance_mod',
                                  callback_data=f'change_appearance_mod={slide_id}'))\
        .row(InlineKeyboardButton(text='change_schedule_set',
                                  callback_data=f'change_schedule_set={slide_id}'))\
        .row(InlineKeyboardButton(text='change_schedule_priority',
                                  callback_data=f'change_schedule_priority={slide_id}'))\
        .row(InlineKeyboardButton(text='change_header',
                                  callback_data=f'change_header={slide_id}'))\
        .row(InlineKeyboardButton(text='change_buttons',
                                  callback_data=f'change_buttons={slide_id}'))


def button_change_kb(slide_id: int, row_num: int, row_pos: int):
    return InlineKeyboardMarkup()\
        .row(InlineKeyboardButton(text='chg_button_row_num',
                                  callback_data=f'chg_button_row_num={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_row_pos',
                                  callback_data=f'chg_button_row_pos={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_slide_id',
                                  callback_data=f'chg_button_slide_id={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_slide_link',
                                  callback_data=f'chg_button_slide_link={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_name',
                                  callback_data=f'chg_button_name={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_url',
                                  callback_data=f'chg_button_url={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_modifier',
                                  callback_data=f'chg_button_modifier={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='chg_button_appearance_mod',
                                  callback_data=f'chg_button_appearance_mod={slide_id},{row_num},{row_pos}'))\
        .row(InlineKeyboardButton(text='delete_button',
                                  callback_data=f'delete_button={slide_id},{row_num},{row_pos}'))


class FSMSlideChange(StatesGroup):
    slide_change_media = State()
    slide_change_message = State()
    slide_change_bot_id = State()
    slide_change_modifier = State()
    slide_change_appearance_mod = State()
    slide_change_schedule_set = State()
    slide_change_schedule_priority = State()
    slide_change_header = State()
    slide_change_buttons = State()
    button_change_row_num = State()
    button_change_row_pos = State()
    button_change_slide_id = State()
    button_change_slide_link = State()
    button_change_name = State()
    button_change_url = State()
    button_change_modifier = State()
    button_change_appearance_mod = State()


class FSMAdminDatabase(StatesGroup):
    start_db_command = State()


async def cm_cancel(message: types.Message, state: FSMContext):
    if message.from_user.id in admin_ids:
        cur_state = await state.get_state()
        if cur_state is None:
            return
        await state.finish()
        await message.reply('Отменено')


async def cm_start_db_command(message: types.Message):
    if message.from_user.id in admin_ids:
        await FSMAdminDatabase.start_db_command.set()
        await message.answer('```\nSQL command (qq - exit):\n```', parse_mode='MarkdownV2')
    await delete_message(message)


async def commit_db_command(message: types.Message, upload_type=None):
    low_text = message.text.lower()
    try:
        cur.execute(message.text)
        if low_text[:6] == 'select' or low_text[:4] == 'with':
            res = cur.fetchall()

            if upload_type == 'excel':
                try:
                    workbook = Workbook()
                    sheet = workbook.active
                    # Запись заголовков столбцов
                    columns = [desc[0] for desc in cur.description]
                    for col_index, column_title in enumerate(columns, 1):
                        sheet.cell(row=1, column=col_index, value=column_title)
                    # Запись данных
                    for row_index, row in enumerate(res, 2):
                        for col_index, cell_value in enumerate(row, 1):
                            sheet.cell(row=row_index, column=col_index, value=row[cell_value])
                    # Сохранение файла Excel
                    workbook.save(filename="data.xlsx")
                    file = InputFile('data.xlsx')
                    await message.answer_document(file)
                except Exception as err:
                    await group_msg(f"ERR commit_db_command.excel - {type(err).__name__}")

            else:  # Вывод таблицы в чат
                MAX_MESSAGE_LENGTH = 4000
                MAX_LINE_LENGTH = 142

                # Заголовки таблицы
                wide_matrix = []
                column_name_width = []
                column_names = []
                for column_name, column_val in res[0].items():
                    column_name_width.append(len(str(column_name)))
                    column_names.append(column_name)
                    wide_matrix.append([])

                # Ширина столбцов каждой строки. 0 - если выходит за лимиты и требуется перенос строки
                for row in res:
                    cur_col_num = 0
                    cur_line_length = 0
                    for column_name, column_val in row.items():
                        column_length = len(str(column_val))
                        wide_matrix[cur_col_num].append(column_length)
                        cur_line_length += column_length + 1
                        cur_col_num += 1

                # Рассылка
                cur_row = 0
                first_msg_row = 0
                name_wide_matrix = []
                for col_width in column_name_width:
                    name_wide_matrix.append([col_width])

                async def print_row(p_row):
                    final_msg = ''
                    for col_name, col_val in p_row.items():
                        final_msg += f"<b>{col_name}</b>: {col_val}\n"
                    await message.answer(final_msg, parse_mode='HTML')

                final_message = ''
                while cur_row < len(wide_matrix[0]):
                    cur_message = ''

                    # Ширина столбцов в отрезке
                    local_wide_matrix = list()
                    for i in range(len(name_wide_matrix)):
                        local_wide_matrix.append(name_wide_matrix[i][:])
                    for col_id in range(len(local_wide_matrix)):
                        local_wide_matrix[col_id] += wide_matrix[col_id][first_msg_row:cur_row+1]

                    # Максимальная ширина столбца в отрезке:
                    max_local_wide_matrix = list(map(max, local_wide_matrix))

                    # Заголовки:
                    for itr in range(len(column_names)):
                        cur_message += column_names[itr].ljust(max_local_wide_matrix[itr]+1)
                    cur_message = cur_message.rstrip()+'\n'

                    # Строки:
                    for itr in range(first_msg_row, cur_row+1):
                        cur_row_text = ''
                        for col_number in range(len(column_names)):
                            text_addition = str(res[itr][column_names[col_number]])
                            if (len(cur_row_text) - cur_row_text.rfind('\n') - 1) + len(text_addition) <= MAX_LINE_LENGTH:
                                cur_row_text += text_addition.ljust(max_local_wide_matrix[col_number]+1)
                            else:
                                cur_row_text = cur_row_text.lstrip()+'\n'+text_addition+' '
                        cur_message += cur_row_text.lstrip()+'\n'

                    # Цикл
                    if len(cur_message) <= MAX_MESSAGE_LENGTH:
                        final_message = cur_message
                        if cur_row + 1 == len(wide_matrix[0]):
                            if cur_row == first_msg_row:
                                await print_row(res[cur_row])
                            else:
                                await message.answer(f'```result\n{final_message}```', parse_mode='MarkdownV2')
                        cur_row += 1
                    else:
                        if final_message:
                            if cur_row - 1 == first_msg_row:
                                await print_row(res[first_msg_row])
                            else:
                                await message.answer(f'```result\n{final_message}```', parse_mode='MarkdownV2')
                        else:
                            await print_row(res[first_msg_row])
                            cur_row += 1
                        final_message = ''
                        first_msg_row = cur_row
        else:
            base.commit()
            await message.answer('```\ncommited\n```', parse_mode='MarkdownV2')
    except Exception as err:
        await message.answer("```\nSQL Error: " + str(err) + '\n```', parse_mode='MarkdownV2')
        cur.execute('rollback')


async def cm_commit_db_command(message: types.Message, state: FSMContext):
    if message.from_user.id in admin_ids:
        low_text = message.text.lower()
        if low_text != 'qq':
            await commit_db_command(message)
        else:
            await state.finish()
            await delete_message(message)


async def test_photo_post(message: types.Message):
    if message.from_user.id in admin_ids:
        try:
            await message.answer_photo(message.text.split(' ')[1], caption=message.text.split(' ')[1])
        except WrongFileIdentifier:
            await group_msg('WrongFileIdentifier: Медиафайл другого бота')
        except Exception as err:
            await group_msg(err)
    await delete_message(message)


async def test_video_post(message: types.Message):
    if message.from_user.id in admin_ids:
        try:
            await message.answer_video(message.text.split(' ')[1], caption=message.text.split(' ')[1])
        except WrongFileIdentifier:
            await group_msg('WrongFileIdentifier: Медиафайл другого бота')
        except Exception as err:
            await group_msg(err)
    await delete_message(message)


async def test_file_post(message: types.Message):
    if message.from_user.id in admin_ids:
        try:
            await message.answer_document(message.text.split(' ')[1], caption=message.text.split(' ')[1])
        except WrongFileIdentifier:
            await group_msg('WrongFileIdentifier: Медиафайл другого бота')
        except Exception as err:
            await group_msg(err)
    await delete_message(message)


async def test_voice_post(message: types.Message):
    if message.from_user.id in admin_ids:
        try:
            await message.answer_voice(message.text.split(' ')[1], caption=message.text.split(' ')[1])
        except WrongFileIdentifier:
            await group_msg('WrongFileIdentifier: Медиафайл другого бота')
        except Exception as err:
            await group_msg(err)
    await delete_message(message)


async def test_video_note_post(message: types.Message):
    if message.from_user.id in admin_ids:
        try:
            await message.answer_voice(message.text.split(' ')[1], caption=message.text.split(' ')[1])
        except WrongFileIdentifier:
            await group_msg('WrongFileIdentifier: Медиафайл другого бота')
        except Exception as err:
            await group_msg(err)
    await delete_message(message)


async def test_gif_post(message: types.Message):
    if message.from_user.id in admin_ids:
        try:
            await message.answer_animation(message.text.split(' ')[1], caption=message.text.split(' ')[1])
        except WrongFileIdentifier:
            await group_msg('WrongFileIdentifier: Медиафайл другого бота')
        except Exception as err:
            await group_msg(err)
    await delete_message(message)


async def test_raw_slide_post(message: types.Message):
    if message.from_user.id in admin_ids:
        cmd_list = message.text.split(' ')
        if len(cmd_list) == 2 and cmd_list[1].isdigit():
            slide_id = cmd_list[1]
            try:
                slide = get_slide_deprecated(int(slide_id), bot_id)
                try:
                    await construct_slide(slide, message)
                except SlideError:
                    await message.answer("Что-то пошло не так")
            except ValueError:
                await group_msg('Неверный номер слайда')
            except Exception as err:
                await group_msg(f"test_raw_slide_post - {err}")
        else:
            await group_msg('Неверный номер слайда')
    await delete_message(message)


async def test_slide_menu(slide_id: str, user_id=admin_ids[0]):
    try:
        if slide_id == 'new':
            db_execute(f"insert into md_slides (bot_id, message) values ('{bot_id}', 'empty')")
            slide_id = db_execute(f"select max(id) as id from md_slides")[0]["id"]
            await test_slide_menu(slide_id, user_id)
        elif type(slide_id) is int or slide_id.isdigit():
            slide = get_slide_deprecated(int(slide_id), bot_id)
            for usr in admin_ids:
                db_execute(f"delete from ft_user_activity where usr_id = {usr};")
            await group_msg(f"slide = {slide['id']}\n"
                            f"header = {slide['header']}\n"
                            f"media_id = {slide['media_id']}\n"
                            f"type = {slide['type']}\n"
                            f"message = {slide['message']}\n"
                            f"bot_id = {slide['bot_id']}\n"
                            f"modifier = {slide['modifier']}\n"
                            f"appearance_mod = {slide['appearance_mod']}\n"
                            f"schedule_set = {slide['schedule_set']}\n"
                            f"schedule_priority = {slide['schedule_priority']}",
                            keyboard=slide_change_kb(slide['id']),
                            parse_mode=None)
            try:
                await construct_slide_from_message(slide, user_id=user_id)
            except SlideError:
                await group_msg("Что-то пошло не так")
            for usr in admin_ids:
                db_execute(f"delete from ft_user_activity where usr_id = {usr}; "
                           f"delete from ft_schedule where usr_id = {usr}; ")
    except ValueError:
        await group_msg('Неверный номер слайда')
    except Exception as err:
        await group_msg(f"test_slide_post - {err} / {slide_id}")


async def test_slide_post(message: types.Message):
    if message.from_user.id in admin_ids:
        cmd_list = re.split(r'[ _]', message.text)
        if len(cmd_list) == 2:
            slide_id = cmd_list[1]
            await test_slide_menu(slide_id, message.from_user.id)
        else:
            await group_msg('Неверный номер слайда')
    await delete_message(message)


@dp.callback_query_handler(text_startswith="test_slide_post", state=FSMSlideChange.slide_change_buttons)
async def test_slide_post_callback(callback: types.CallbackQuery, state: FSMContext):
    slide_id = callback.data.split('=')[1].split(',')[0]
    await state.finish()
    await test_slide_menu(slide_id, callback.from_user.id)
    await callback.answer()


@dp.callback_query_handler(text_startswith="change_message")
async def change_slide_message(callback: types.CallbackQuery, state: FSMContext):
    await group_msg('Type new message below:')
    await FSMSlideChange.slide_change_message.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_message(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    db_execute(f"update md_slides set message = {insert_value} where id = {slide_id}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_media")
async def change_slide_media(callback: types.CallbackQuery, state: FSMContext):
    await group_msg('Type new media below:')
    await FSMSlideChange.slide_change_media.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_media(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    media_id = None
    if message.content_type == 'text' and message.text == 'null':
        db_execute(f"update md_slides set media_id = null where id = {slide_id}")
    else:
        if message.content_type == 'photo':
            media_id = message.photo[0].file_id
        elif message.content_type == 'video':
            media_id = message.video.file_id
        elif message.content_type == 'document':
            media_id = message.document.file_id
        elif message.content_type == 'voice':
            media_id = message.voice.file_id
        elif message.content_type == 'video_note':
            media_id = message.video_note.file_id
        elif message.content_type == 'animation':
            media_id = message.animation.file_id
        if media_id:
            select = db_execute(f"select * from md_media where media_id = '{media_id}' and type = '{message.content_type}'")
            if not select:
                insert_media(media_id, message.content_type)  # db_execute(f"insert into md_media (type, media_id) values ('{message.content_type}', '{media_id}')")
            db_execute(f"update md_slides set media_id = (select max(id) from md_media where media_id = '{media_id}') "
                       f"where id = {slide_id}")
        else:
            await group_msg('receive_changed_media - пустой media_id')
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_bot_id")
async def change_bot_id(callback: types.CallbackQuery, state: FSMContext):
    select = db_execute(f"select distinct name from md_bots")
    names = ''
    for row in select:
        names = names + f"{row['name']}\n"
    await group_msg(f"Type new bot_id below:\n"
                    f"{names}")
    await FSMSlideChange.slide_change_bot_id.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_bot_id(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    try:
        db_execute(f"update md_slides set bot_id = '{message.text}' where id = {slide_id}")
    except Exception as err:
        await group_msg(f"receive_changed_bot_id - {err}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_modifier")
async def change_modifier(callback: types.CallbackQuery, state: FSMContext):
    await group_msg('Type new bot_id below:\n'
                    "Modifiers:\n"
                    "empty - слайд не покажут\n"
                    "payment_start - для реквизитов оплаты и выборе курса\n"
                    "redirect=1 - перенаправляет на указаный слайд вместо показа текущего\n"
                    "quest_multiple - для вопросов опроса с множественными вариантами ответа\n"
                    "Уникальные для каждого курса: \n"
                    "course_reject_1 - об отказе оплаты курса\n"
                    "course_paid_1 - о подтверждении оплаты курса\n"
                    "course_price_1 - при выборе курса для оплаты\n"
                    "coupon_start_1 - при выдаче купона\n"
                    "Уникальные для каждого бота: \n"
                    "correct_payment_file - при принятии корректного файла оплаты\n"
                    "incorrect_payment_file - при принятии некорректного файла оплаты\n"
                    "start - стартовый")
    await FSMSlideChange.slide_change_modifier.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_modifier(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        if message.text in ['start', 'correct_payment_file', 'incorrect_payment_file'] or message.text.startswith(('course_reject_', 'course_paid_', 'course_price_', 'coupon_start_')):
            db_execute(f"update md_slides set modifier = null where id = (select id from md_slides where modifier = '{message.text}' and bot_id = '{bot_id}')")
        db_execute(f"update md_slides set modifier = {insert_value} where id = {slide_id}")
    except Exception as err:
        await group_msg(f"receive_changed_modifier - {err}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_appearance_mod")
async def change_appearance_mod(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new appearance_mod below:\n"
                    "coupon_expire=1\n"
                    "course_paid=1\n"
                    "course_not_paid=1\n"
                    "disappear_any=1,2\n"
                    "disappear_all=1,2\n"
                    "appear_any=1,2\n"
                    "appear_all=1,2")
    await FSMSlideChange.slide_change_appearance_mod.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_appearance_mod(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_slides set appearance_mod = {insert_value} where id = {slide_id}")
    except Exception as err:
        await group_msg(f"receive_changed_appearance_mod - {err}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_schedule_set")
async def change_schedule_set(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new schedule_set below:\n"
                    "weekday=0,week=0,hour=13,minute=30,slide=2;day=2,hour=10,minute=20,slide=3;minutes=30,slide=4")
    await FSMSlideChange.slide_change_schedule_set.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_schedule_set(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_slides set schedule_set = {insert_value} where id = {slide_id}")
    except Exception as err:
        await group_msg(f"receive_changed_schedule_set - {err}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_schedule_priority")
async def change_schedule_priority(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new schedule_priority below:")
    await FSMSlideChange.slide_change_schedule_priority.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_schedule_priority(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_slides set schedule_priority = {insert_value} where id = {slide_id}")
    except Exception as err:
        await group_msg(f"receive_changed_schedule_priority - {err}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


@dp.callback_query_handler(text_startswith="change_header")
async def change_header(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new header below:")
    await FSMSlideChange.slide_change_header.set()
    async with state.proxy() as data:
        data['slide_id'] = int(callback.data.split('=')[1])
    await callback.answer()


async def receive_changed_header(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_slides set header = {insert_value} where id = {slide_id}")
    except Exception as err:
        await group_msg(f"receive_changed_header - {err}")
    message.text = f'slide {slide_id}'
    await state.finish()
    await test_slide_post(message)


async def get_test_keyboard(slide_id):
    keyboard_select = get_keyboard(slide_id)
    for button in keyboard_select:
        button["name"] = f"{button['row_num']}.{button['row_pos']}.{button['name']}"
        button["modifier"] = f"change_btn={slide_id},{button['row_num']},{button['row_pos']}"
    keyboard = await construct_keyboard_from_select(keyboard_select)
    if keyboard is None:
        keyboard = InlineKeyboardMarkup()
    keyboard.row(InlineKeyboardButton(text="Create new button", callback_data=f'new_btn={slide_id}'))
    keyboard.row(InlineKeyboardButton(text="Return", callback_data=f'test_slide_post={slide_id}'))
    return keyboard


@dp.callback_query_handler(text_startswith="change_buttons")
async def change_buttons_menu(value, state: FSMContext):
    if type(value) is types.CallbackQuery:
        slide_id = int(value.data.split('=')[1].split(',')[0])
    elif type(value) is int:
        slide_id = value
    else:
        raise TypeError
    await group_msg("Сhoose button below:", keyboard=await get_test_keyboard(slide_id=slide_id))
    await FSMSlideChange.slide_change_buttons.set()
    async with state.proxy() as data:
        data['slide_id'] = slide_id
    if type(value) is types.CallbackQuery:
        await value.answer()


@dp.callback_query_handler(text_startswith="new_btn", state=FSMSlideChange.slide_change_buttons)
async def create_button(callback: types.CallbackQuery, state: FSMContext):
    create_new_button(int(callback.data.split('=')[1]))
    await change_buttons_menu(callback, state)


@dp.callback_query_handler(text_startswith="change_btn", state=FSMSlideChange.slide_change_buttons)
async def change_button_menu(callback: types.CallbackQuery, state: FSMContext):
    info = callback.data.split('=')[1].split(',')
    button = get_button_by_ids(slide_id=info[0], row_num=info[1], row_pos=info[2])
    await group_msg(f"row_num = {button['row_num']}\n"
                    f"row_pos = {button['row_pos']}\n"
                    f"slide_id = {button['slide_id']}\n"
                    f"slide_link = {button['slide_link']}\n"
                    f"name = {button['name']}\n"
                    f"url = {button['url']}\n"
                    f"modifier = {button['modifier']}\n"
                    f"appearance_mod = {button['appearance_mod']}",
                    keyboard=button_change_kb(slide_id=int(info[0]), row_num=int(info[1]), row_pos=int(info[2])),
                    parse_mode=None)
    await FSMSlideChange.slide_change_buttons.set()
    async with state.proxy() as data:
        data['slide_id'] = info[0]
    await callback.answer()


@dp.callback_query_handler(text_startswith="delete_button", state=FSMSlideChange.slide_change_buttons)
async def delete_button(callback: types.CallbackQuery, state: FSMContext):
    info = callback.data.split('=')[1].split(',')
    delete_button_by_ids(slide_id=info[0], row_num=info[1], row_pos=info[2])
    await change_buttons_menu(callback, state)
    await callback.answer()


@dp.callback_query_handler(text_startswith="chg_button_row_num", state=FSMSlideChange.slide_change_buttons)
async def chg_button_row_num(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new row_num below:")
    await FSMSlideChange.button_change_row_num.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_row_num(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set row_num = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_row_num - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_row_pos", state=FSMSlideChange.slide_change_buttons)
async def chg_button_row_pos(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new row_pos below:")
    await FSMSlideChange.button_change_row_pos.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_row_pos(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set row_pos = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_row_pos - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_slide_id", state=FSMSlideChange.slide_change_buttons)
async def chg_button_slide_id(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new slide_id below:")
    await FSMSlideChange.button_change_slide_id.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_slide_id(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set slide_id = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_slide_id - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_slide_link", state=FSMSlideChange.slide_change_buttons)
async def chg_button_slide_link(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new slide_link below:")
    await FSMSlideChange.button_change_slide_link.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_slide_link(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set slide_link = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_slide_link - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_name", state=FSMSlideChange.slide_change_buttons)
async def chg_button_name(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new name below:")
    await FSMSlideChange.button_change_name.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_name(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set name = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_name - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_url", state=FSMSlideChange.slide_change_buttons)
async def chg_button_url(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("Type new url below:")
    await FSMSlideChange.button_change_url.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_url(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set url = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_url - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_modifier", state=FSMSlideChange.slide_change_buttons)
async def chg_button_modifier(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("merch_pay='merch_id'\n"
                    "paycheck='course_id','accept_slide_id','decline_slide_id'\n"
                    "payment_refuse\n"
                    "questionnaire='questionnaire_id','end_slide_id'\n"
                    "Type new modifier below:")
    await FSMSlideChange.button_change_modifier.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_modifier(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set modifier = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_modifier - {err}")
    await change_buttons_menu(slide_id, state)


@dp.callback_query_handler(text_startswith="chg_button_appearance_mod", state=FSMSlideChange.slide_change_buttons)
async def chg_button_appearance_mod(callback: types.CallbackQuery, state: FSMContext):
    await group_msg("coupon_expire=1\n"
                    "course_paid=1\n"
                    "course_not_paid=1\n"
                    "disappear_any=1,2\n"
                    "disappear_all=1,2\n"
                    "appear_any=1,2\n"
                    "appear_all=1,2\n"
                    "Type new appearance_mod below:")
    await FSMSlideChange.button_change_appearance_mod.set()
    info = callback.data.split('=')[1].split(',')
    async with state.proxy() as data:
        data['slide_id'] = int(info[0])
        data['row_num'] = int(info[1])
        data['row_pos'] = int(info[2])
    await callback.answer()


async def receive_chg_button_appearance_mod(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        slide_id = data['slide_id']
        row_num = data['row_num']
        row_pos = data['row_pos']
    insert_value = message.text
    if message.text != 'null':
        insert_value = f"'{insert_value}'"
    try:
        db_execute(f"update md_buttons set appearance_mod = {insert_value} "
                   f"where slide_id = {slide_id} and row_num = {row_num} and row_pos = {row_pos}")
    except Exception as err:
        await group_msg(f"receive_chg_button_appearance_mod - {err}")
    await change_buttons_menu(slide_id, state)


async def obj_testing(message: types.Message):
    if message.from_user.id in admin_ids:
        test_id = int(message.text.split(' ')[1])
        entity = entities.Merchandise(test_id)

        courses_str = ''
        for course in entity.courses:
            courses_str += f"course_id: {course.id} - {type(course.id).__name__}\n"\
                           f"course_name: {course.name} - {type(course.name).__name__}\n"\
                           f"course_bot_id: {course.bot_id} - {type(course.bot_id).__name__}\n"

        await group_msg(f"id: {entity.id} - {type(entity.id).__name__}\n"
                        f"name: {entity.name} - {type(entity.name).__name__}\n"
                        f"price: {entity.price} - {type(entity.price).__name__}\n"
                        f"courses:\n{courses_str}")


def register_handlers_admin():  # Порядок Важен!
    dp.register_message_handler(obj_testing, Text(startswith='/tst'), state='*')  # УДАЛИТЬ ПОСЛЕ ТЕСТИРОВАНИЯ - ЦЕ ХУIТА
    dp.register_message_handler(cm_cancel, state='*', commands='отмена')
    dp.register_message_handler(cm_cancel, Text(equals='/отмена', ignore_case=True), state='*')
    dp.register_message_handler(receive_changed_message, state=FSMSlideChange.slide_change_message)
    dp.register_message_handler(receive_changed_media, state=FSMSlideChange.slide_change_media,
                                content_types=['text', 'photo', 'video', 'document', 'voice', 'video_note', 'animation'])
    dp.register_message_handler(receive_changed_bot_id, state=FSMSlideChange.slide_change_bot_id)
    dp.register_message_handler(receive_changed_modifier, state=FSMSlideChange.slide_change_modifier)
    dp.register_message_handler(receive_changed_appearance_mod, state=FSMSlideChange.slide_change_appearance_mod)
    dp.register_message_handler(receive_changed_schedule_set, state=FSMSlideChange.slide_change_schedule_set)
    dp.register_message_handler(receive_changed_schedule_priority, state=FSMSlideChange.slide_change_schedule_priority)
    dp.register_message_handler(receive_changed_header, state=FSMSlideChange.slide_change_header)
    dp.register_message_handler(receive_chg_button_row_num, state=FSMSlideChange.button_change_row_num)
    dp.register_message_handler(receive_chg_button_row_pos, state=FSMSlideChange.button_change_row_pos)
    dp.register_message_handler(receive_chg_button_slide_id, state=FSMSlideChange.button_change_slide_id)
    dp.register_message_handler(receive_chg_button_slide_link, state=FSMSlideChange.button_change_slide_link)
    dp.register_message_handler(receive_chg_button_name, state=FSMSlideChange.button_change_name)
    dp.register_message_handler(receive_chg_button_url, state=FSMSlideChange.button_change_url)
    dp.register_message_handler(receive_chg_button_modifier, state=FSMSlideChange.button_change_modifier)
    dp.register_message_handler(receive_chg_button_appearance_mod, state=FSMSlideChange.button_change_appearance_mod)
    dp.register_message_handler(cm_start_db_command, commands='dbc', state=None)
    dp.register_message_handler(cm_commit_db_command, state=FSMAdminDatabase.start_db_command)
    dp.register_message_handler(test_photo_post, commands=['p'])
    dp.register_message_handler(test_video_post, commands=['v'])
    dp.register_message_handler(test_file_post, commands=['d'])
    dp.register_message_handler(test_voice_post, commands=['a'])
    dp.register_message_handler(test_video_note_post, commands=['vn'])
    dp.register_message_handler(test_gif_post, commands=['g'])
    dp.register_message_handler(test_slide_post, Text(startswith='/slide', ignore_case=True))  # commands=['slide'])
    dp.register_message_handler(test_raw_slide_post, commands=['raw'])
